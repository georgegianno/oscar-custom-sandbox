from django.utils.translation import gettext_lazy as _
from django.http import HttpResponse
from oscar.apps.order.models import Guest
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from io import BytesIO
from django.db.models import Count
from operator import attrgetter
from itertools import chain
from django.db.models import Sum
from oscar.core.loading import get_model
from oscar.core.compat import get_user_model
from django.utils import timezone


User = get_user_model()
Order = get_model('order', 'Order')

def export_customers(request, customers):
    wb = Workbook()
    sheet = wb.active
    headers = ['Email','Name','Is user','Date joined', 'Orders count','Orders total value']
    if customers == 'customers':
        del headers[2]
        filename = 'customers_export.xlsx'
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            sheet["%s1" % col_letter] = header
        col_num=2 
        for customer in User.objects.annotate(count=Count('orders')).order_by('-count'):
            sheet["A%d" % col_num] = customer.email if customer.email else '-'
            sheet["B%d" % col_num] = str(customer.first_name)+str(customer.last_name) if customer.first_name or customer.last_name else '-'  
            sheet["C%d" % col_num] = timezone.make_naive(customer.date_joined) if customer.date_joined else '-'

            sheet["D%d" % col_num] = customer.orders.count() if customer.orders else '0'
            sheet["E%d" % col_num] = float(customer.orders.filter(status="Completed"). \
                aggregate(Sum('total_incl_tax')).get('total_incl_tax__sum')) if customer.orders.filter(status="Completed") else '0'
            col_num += 1
    if customers == 'guests':
        del headers[2]
        filename = 'guests_export.xlsx'
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            sheet["%s1" % col_letter] = header
        col_num=2 
        for guest in Guest.objects.annotate(count=Count('orders')).order_by('-count'):
            sheet["A%d" % col_num] = guest.email if guest.email else '-'
            sheet["B%d" % col_num] = guest.name if guest.name else '-'  
            sheet["D%d" % col_num] = timezone.make_naive(guest.orders.order_by('date_placed').first().date_placed) if guest.orders.exists() else '-'
            sheet["D%d" % col_num] = guest.orders.count() if guest.orders else '0'
            sheet["E%d" % col_num] = float(guest.orders.filter(status="Completed"). \
                aggregate(Sum('total_incl_tax')).get('total_incl_tax__sum')) if guest.orders.filter(status="Completed") else '0'
            col_num += 1
    if customers == 'all':
        filename = 'all_export.xlsx'
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            sheet["%s1" % col_letter] = header
        col_num=2 
        for user in sorted(list(chain(Guest.objects.exclude(email__in=User.objects.values_list('email', flat=True)). \
                annotate(count=Count('orders')), User.objects.annotate(count=Count('orders')))), key=attrgetter('count'), reverse=True):
            orders =  Order.objects.filter(status='Completed').filter(guest_email=user.email)
            sheet["A%d" % col_num] = user.email if user.email else '-'
            sheet["F%d" % col_num] = float(orders.aggregate(Sum('total_incl_tax')).get('total_incl_tax__sum')) if orders else '0'
            if isinstance(user, Guest):
                sheet["B%d" % col_num] = user.name if user.name else '-' 
                sheet["C%d" % col_num] = 'no'  
                sheet["D%d" % col_num] = timezone.make_naive(user.orders.order_by('date_placed').first().date_placed) if user.orders.exists() else '-'
                sheet["E%d" % col_num] = user.orders.count() if user.orders else 0
            else:
                orders_count = int(user.orders.count()) if user.orders else 0
                was_guest = Guest.objects.filter(email=user.email)
                if was_guest:
                    as_guest = was_guest.first()
                    orders_count += as_guest.orders.count()
                sheet["B%d" % col_num] = str(user.first_name)+str(user.last_name) if user.first_name or user.last_name else '-'  
                sheet["C%d" % col_num] = 'yes'  
                sheet["D%d" % col_num] = user.date_joined
                sheet["E%d" % col_num] = orders_count 
            col_num += 1
    output = BytesIO()
    wb.save(output)
    output.seek(0)  
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename={}'.format(filename)
    response.write(output.getvalue())
    
    return response